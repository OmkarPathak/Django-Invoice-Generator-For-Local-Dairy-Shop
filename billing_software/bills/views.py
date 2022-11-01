import encodings
from django.shortcuts import render, redirect
from .models import Upload, UploadForm, SingleBillForm, Rate, RateForm
from django.contrib import messages
from openpyxl import load_workbook
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from xlsxwriter.workbook import Workbook
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string, get_template
from xhtml2pdf import pisa
import io
import os, datetime
import random, string
from pathlib import Path

def get_random_text():
    return ''.join(random.sample(string.ascii_letters, 6))

today = datetime.datetime.now()

def generate_single_bill(request):
    rates = Rate.objects.get(id=1)
    if request.method == 'POST':
        pass
    form = SingleBillForm()
    template = get_template('single_bill.html')
    html = template.render({'form': form, 'rates': rates})
    # rendered = render_to_string('single_bill.html', {'form': form, 'rates': rates})
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode('utf-8')), result, encoding='UTF-8')
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    # return HttpResponse(html)
    # pdf = PDF()
    # pdf.set_font_size(16)
    # pdf.add_page()
    # pdf.write_html(rendered, table_line_separators=True)
    # response = HttpResponse(pdf, content_type='application/pdf')
    # response['Content-Disposition'] = "attachment; filename='file_name.pdf'"
    # return response
    # return render(request, 'single_bill.html', {'form': form, 'rates': rates})

def rates(request):
    rates = Rate.objects.get(id=1)
    if request.method == 'POST':
        form = RateForm(request.POST, instance=rates)
        form.save()
        messages.success(request, 'Rates updated')
    form = RateForm(initial={
        # 'cow_milk_rate': rates.cow_milk_rate,
        # 'milk_rate': rates.milk_rate,
        # 'ghee_rate': rates.ghee_rate,
        # 'dahi_rate': rates.dahi_rate
        'rate_updated': rates.rate_updated
    })
    return render(request, 'rates.html', {'form': form})

def upload(request):
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            request.session.clear()
            messages.success(request, 'File uploaded!')
            wb = load_workbook(request.FILES.get('excel_file'), data_only=True)
            # print(wb.sheetnames)
            excel_data = list()
            # # iterating over the rows and
            # # getting value from each cell in row
            for row in wb['Sheet1'].iter_rows():
                row_data = list()
                # print(row)
                if not row[0].value and row[5].value is None:
                    continue
                for cell in row:
                    # print(cell.value)
                    row_data.append(str(cell.value))

                excel_data.append(row_data)
                
            request.session['excel_data'] = excel_data[1:]
            return redirect('upload')
    else:
        form = UploadForm()
        if request.session.get('excel_data'):
            excel_data = request.session.get('excel_data')
            for i in range(len(excel_data)):
                excel_data[i] = excel_data[i][0:2] + excel_data[i][2:9]
            # request.session.clear()
            page = request.GET.get('page', 1)

            # paginate the data
            paginator = Paginator(excel_data, 50)
            try:
                excel_data = paginator.page(page)
            except PageNotAnInteger:
                excel_data = paginator.page(1)
            except EmptyPage:
                excel_data = paginator.page(paginator.num_pages)

            return render(request, 'upload.html', {'form': form, 'excel_data': excel_data})
    return render(request, 'upload.html', {'form': form})

def generate(request):
    output = io.BytesIO()
    excel_data = request.session.get('excel_data')
    # print(excel_data)
    rates = Rate.objects.get()
    # print('COW milk:', rates.cow_milk_rate, rates.milk_rate)
    g_rate      = excel_data[0][10]
    m_rate      = excel_data[0][11]
    dahi_rate   = excel_data[0][12]
    ghee_rate   = excel_data[0][13]
    bill_for    = excel_data[0][14]
    bill_gen    = excel_data[0][15]

    for i in range(len(excel_data)):
        excel_data[i] = excel_data[i][0:2] + excel_data[i][2:9]

    file_name = today.strftime('%B') + '_' + today.strftime('%Y')
    
    # Check if the dir already exists
    # try:
    #     os.makedirs(os.path.join(settings.MEDIA_ROOT, today.strftime('bills/%Y/%B/')))
    # except FileExistsError:
    #     pass

    # check_file = Path(os.path.join(settings.MEDIA_ROOT, today.strftime('bills/%Y/%B/'), file_name + '.xlsx'))

    # # check if file already exists
    # if check_file.exists():
    #     book = Workbook(os.path.join(settings.MEDIA_ROOT, today.strftime('bills/%Y/%B/'), file_name + '_' + get_random_text() + '.xlsx'))
    #     sheet = book.add_worksheet('Sheet1')
    # else:
    #     book = Workbook(os.path.join(settings.MEDIA_ROOT, today.strftime('bills/%Y/%B/'), file_name + '.xlsx'))
    #     sheet = book.add_worksheet('Sheet1')
    book = Workbook(output, {'in_memory': True})
    sheet = book.add_worksheet('Sheet1')
    # default cell format to size 10 
    # book.formats[0].set_font_size(12)

    for col in range(1000):
        sheet.set_column(col, col, 8)
        sheet.set_row(col, 25)

    merge_format = book.add_format({
        'bold': 3,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 12
    })

    custom_font = book.add_format({
        'bold': 3,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 12
    })

    try:
        custom_font.set_font_name('Shivaji01')
    except KeyError:
        custom_font.set_font_name('Arial')
        
    sign = book.add_format({
        'bold': 3,
        'border': 1,
        'align': 'right',
        'valign': 'vcenter',
        'font_size': 12
    })
    
    heading = book.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 12
    })

    data = book.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 10
    })

    border = book.add_format({'border': 1})
    
    row = 0
    col = 0
    
    for i in range(0, len(excel_data), 4):
        # print(excel_data[i])

        sr_no       = excel_data[i][0]
        name        = excel_data[i][1]

        # A1:B1
        sheet.merge_range(
            'A' + str(row + 1) + ':B' + str(row + 1), 
            'क्र: ' + str(sr_no), 
            data
        )

        # E1:F1
        sheet.merge_range(
            'E' + str(row + 1) + ':F' + str(row + 1), 
            'फोन: 9371079745', 
            data
        )

        # C1:D1
        sheet.merge_range(
            'C' + str(row + 1) + ':D' + str(row + 1), 
            '', 
            merge_format)

        # A2:F2
        sheet.merge_range(
            'A' + str(row + 2) + ':F' + str(row + 2), 
            'मातृछाया दुग्धालय\n 210, कसबा पेठ, पुणे-411011', 
            heading)

        sheet.set_row(1, 30)

        # A3:F3
        sheet.merge_range(
            'A' + str(row + 3) + ':F' + str(row + 3), 
            name, 
            custom_font)

        # A5:C5
        sheet.merge_range(
            'A' + str(row + 5) + ':C' + str(row + 5), 
            'तपशील', 
            heading
        )

        # D5
        sheet.write(
            'D'  + str(row + 5), 
            'प्रमाण', 
            heading
        )

        # E5
        sheet.write(
            'E' + str(row + 5) , 
            'दर', 
            heading
        )

        # F5
        sheet.write(
            'F' + str(row + 5), 
            'रक्कम', 
            heading
        )

        # A4:C4
        sheet.merge_range(
            'A' + str(row + 4) + ':C' + str(row + 4), 
            datetime.datetime.strptime(bill_for, "%Y-%m-%d %H:%M:%S").strftime("%B-%Y"), 
            merge_format)

        # D4:F4
        sheet.merge_range(
            'D' + str(row + 4) + ':F' + str(row + 4), 
            'तारीख: ' + str(datetime.datetime.strptime(bill_gen, "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")), 
            data)

        quantity        = excel_data[i][2]
        dahi            = excel_data[i][3]
        ghee            = excel_data[i][4]
        milk_type       = excel_data[i][5]
        amount          = excel_data[i][6]
        previous_blnc   = excel_data[i][7]
        total           = excel_data[i][8]

        quantity1   = excel_data[i + 1][2]
        milk_type1  = excel_data[i + 1][5]
        amount1     = excel_data[i + 1][6]
        
        # A6:C6
        sheet.merge_range(
            'A' + str(row + 6) + ':C' + str(row + 6), 
            'म्हशींचे दूध', 
            data
        )
        if milk_type == 'M':
            sheet.write('D' + str(row + 6), round(float(quantity), 2), data)
            sheet.write('E' + str(row + 6), str(m_rate) + '/Ltr', data)
            sheet.write('F' + str(row + 6), round(float(amount), 2), data)
        elif milk_type1 == 'M':
            sheet.write('D' + str(row + 6), round(float(quantity1), 2), data)
            sheet.write('E' + str(row + 6), str(m_rate) + '/Ltr', data)
            sheet.write('F' + str(row + 6), round(float(amount1), 2), data)
        else:
            sheet.write_blank('D' + str(row + 6), '', data)
            sheet.write_blank('E' + str(row + 6), '', data)
            sheet.write_blank('F' + str(row + 6), '', data)

        # A7:C7
        sheet.merge_range(
            'A' + str(row + 7) + ':C' + str(row + 7), 
            'गायीचे दूध', 
            data
        )
        if milk_type == 'G':
            sheet.write('D'  + str(row + 7), round(float(quantity), 2), data)
            sheet.write('E' + str(row + 7), str(g_rate) + '/Ltr', data)
            sheet.write('F' + str(row + 7), round(float(amount), 2), data)
        elif milk_type1 == 'G':
            sheet.write('D' + str(row + 7), round(float(quantity1), 2), data)
            sheet.write('E' + str(row + 7), str(g_rate) + '/Ltr', data)
            sheet.write('F' + str(row + 7), round(float(amount1), 2), data)
        else:
            sheet.write_blank('D'  + str(row + 7), '', data)
            sheet.write_blank('E' + str(row + 7), '', data)
            sheet.write_blank('F' + str(row + 7), '', data)

        # A8:C8
        sheet.merge_range(
            'A' + str(row + 8) + ':C' + str(row + 8), 
            'दही', 
            data
        )

        # D8
        sheet.write_blank('D' + str(row + 8), '', data)

        # E8
        sheet.write('E' + str(row + 8), str(dahi_rate) + '/Kg', data)

        # F8
        sheet.write('F' + str(row + 8), dahi, data)

        # A9:C9
        sheet.merge_range(
            'A' + str(row + 9) + ':C' + str(row + 9), 
            'तूप', 
            data
        )

        # D9
        sheet.write_blank('D' + str(row + 9), '', data)

        # E9
        sheet.write('E' + str(row + 9), str(ghee_rate) + '/Kg', data)

        # F9
        sheet.write('F' + str(row + 9), ghee, data)
        # sheet.merge_range('A9:C9', 'मागील शिल्लक', data)
        
        # A10:C10
        sheet.merge_range(
            'A' + str(row + 10) + ':C' + str(row + 10), 
            'मागील बाकी', 
            data
        )

        # A11:C11
        if rates.rate_updated:
            sheet.merge_range(
                'A' + str(row + 11) + ':C' + str(row + 11), 
                'दि. ' + str(rates.rate_updated.strftime('%d/%m/%Y')) + ' पासून', 
                data
            )
        else:
            sheet.merge_range(
                'A' + str(row + 11) + ':C' + str(row + 11), 
                'दि. / / पासून', 
                data
            )

        # A12:C12
        sheet.merge_range(
            'A' + str(row + 12) + ':C' + str(row + 12), 
            'एकूण', 
            data
        )
        
        # D10
        sheet.write_blank('D' + str(row + 10), '', data)

        # E10
        sheet.write_blank('E' + str(row + 10), '', data)

        # F10
        sheet.write('F' + str(row + 10), previous_blnc, data)

        # F11
        sheet.write('F' + str(row + 12), round(float(total), 2), data)

        # D11
        sheet.write_blank('D' + str(row + 11), '', data)

        # E11 (get date here)
        sheet.write('E' + str(row + 11), '', data)

        # F11
        sheet.write_blank('F' + str(row + 11), '', data)

        # F12
        sheet.write('L' + str(row + 12), round(float(total), 2), data)

        # A12:C13
        sheet.merge_range(
            'A' + str(row + 13) + ':F' + str(row + 14), 
            'प्राप्तकर्त्याची स्वाक्षरी', 
            sign
        )

        try:
            #############################################################
            sr_no       = excel_data[i + 2][0]
            name        = excel_data[i + 2][1]

            # G1:H1
            sheet.merge_range(
                'G' + str(row + 1) + ':H' + str(row + 1), 
                'क्र: ' + str(sr_no), 
                data
            )

            # K1:L1
            sheet.merge_range(
                'K' + str(row + 1) + ':L' + str(row + 1), 
                'फोन: 9371079745', 
                data
            )

            # I1:J1
            sheet.merge_range(
                'I' + str(row + 1) + ':J' + str(row + 1), 
                '', 
                merge_format)

            # G2:L2
            sheet.merge_range(
                'G' + str(row + 2) + ':L' + str(row + 2), 
                'मातृछाया दुग्धालय\n 210, कसबा पेठ, पुणे-411011', 
                heading)

            # G3:L3
            sheet.merge_range(
                'G' + str(row + 3) + ':L' + str(row + 3), 
                name, 
                custom_font)

            # G5:I5
            sheet.merge_range(
                'G' + str(row + 5) + ':I' + str(row + 5), 
                'तपशील', 
                heading
            )

            # J5
            sheet.write(
                'J'  + str(row + 5), 
                'प्रमाण', 
                heading
            )

            # K5
            sheet.write(
                'K' + str(row + 5) , 
                'दर', 
                heading
            )

            # L5
            sheet.write(
                'L' + str(row + 5), 
                'रक्कम', 
                heading
            )

            # G4:I4
            sheet.merge_range(
                'G' + str(row + 4) + ':I' + str(row + 4), 
                datetime.datetime.strptime(bill_for, "%Y-%m-%d %H:%M:%S").strftime("%B-%Y"), 
                merge_format)

            # J4:L4
            sheet.merge_range(
                'J' + str(row + 4) + ':L' + str(row + 4), 
                'तारीख: ' + str(datetime.datetime.strptime(bill_gen, "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")), 
                data)

            quantity        = excel_data[i + 2][2]
            dahi            = excel_data[i + 2][3]
            ghee            = excel_data[i + 2][4]
            milk_type       = excel_data[i + 2][5]
            amount          = excel_data[i + 2][6]
            previous_blnc   = excel_data[i + 2][7]
            total           = excel_data[i + 2][8]

            quantity1   = excel_data[i + 3][2]
            milk_type1  = excel_data[i + 3][5]
            amount1     = excel_data[i + 3][6]
            
            # G6:I6
            sheet.merge_range(
                'G' + str(row + 6) + ':I' + str(row + 6), 
                'म्हशींचे दूध', 
                data
            )
            if milk_type == 'M':
                sheet.write('J' + str(row + 6), round(float(quantity), 2), data)
                sheet.write('K' + str(row + 6), str(m_rate) + '/Ltr', data)
                sheet.write('L' + str(row + 6), round(float(amount), 2), data)
            elif milk_type1 == 'M':
                sheet.write('J' + str(row + 6), round(float(quantity1), 2), data)
                sheet.write('K' + str(row + 6), str(m_rate) + '/Ltr', data)
                sheet.write('L' + str(row + 6), round(float(amount1), 2), data)
            else:
                sheet.write_blank('J' + str(row + 6), '', data)
                sheet.write_blank('K' + str(row + 6), '', data)
                sheet.write_blank('L' + str(row + 6), '', data)

            # G7:I7
            sheet.merge_range(
                'G' + str(row + 7) + ':I' + str(row + 7), 
                'गायीचे दूध', 
                data
            )
            if milk_type == 'G':
                sheet.write('J'  + str(row + 7), round(float(quantity), 2), data)
                sheet.write('K' + str(row + 7), str(g_rate) + '/Ltr', data)
                sheet.write('L' + str(row + 7), round(float(amount), 2), data)
            elif milk_type1 == 'G':
                sheet.write('J' + str(row + 7), round(float(quantity1), 2), data)
                sheet.write('K' + str(row + 7), str(g_rate) + '/Ltr', data)
                sheet.write('L' + str(row + 7), round(float(amount1), 2), data)
            else:
                sheet.write_blank('J'  + str(row + 7), '', data)
                sheet.write_blank('K' + str(row + 7), '', data)
                sheet.write_blank('L' + str(row + 7), '', data)

            # G8:I8
            sheet.merge_range(
                'G' + str(row + 8) + ':I' + str(row + 8), 
                'दही', 
                data
            )

            # J8
            sheet.write_blank('J' + str(row + 8), '', data)

            # K8
            sheet.write('K' + str(row + 8), str(dahi_rate) + '/Kg', data)

            # L8
            sheet.write('L' + str(row + 8), dahi, data)

            # G9:I9
            sheet.merge_range(
                'G' + str(row + 9) + ':I' + str(row + 9), 
                'तूप', 
                data
            )

            # J9
            sheet.write_blank('J' + str(row + 9), '', data)

            # K9
            sheet.write('K' + str(row + 9), str(ghee_rate) + '/Kg', data)

            # L9
            sheet.write('L' + str(row + 9), ghee, data)
            # sheet.merge_range('A9:C9', 'मागील शिल्लक', data)

            # G10:I10
            sheet.merge_range(
                'G' + str(row + 10) + ':I' + str(row + 10), 
                'मागील बाकी', 
                data
            )

            # G11:I11
            if rates.rate_updated:
                sheet.merge_range(
                    'G' + str(row + 11) + ':I' + str(row + 11), 
                    'दि. ' + str(rates.rate_updated.strftime('%d/%m/%Y')) + ' पासून', 
                    data
                )
            else:
                sheet.merge_range(
                    'G' + str(row + 11) + ':I' + str(row + 11), 
                    'दि. / / पासून', 
                    data
                )

            # G12:I12
            sheet.merge_range(
                'G' + str(row + 12) + ':I' + str(row + 12), 
                'एकूण', 
                data
            )
            
            # J10
            sheet.write_blank('J' + str(row + 10), '', data)

            # K10
            sheet.write_blank('K' + str(row + 10), '', data)

            # L10
            sheet.write('L' + str(row + 10), previous_blnc, data)

            # J11
            sheet.write_blank('J' + str(row + 11), '', data)

            # K11 (get date here)
            sheet.write_blank('K' + str(row + 11), '', data)

            # L11
            sheet.write_blank('L' + str(row + 11), '', data)

            # L12
            sheet.write('L' + str(row + 12), round(float(total), 2), data)

            # G11:L12
            sheet.merge_range(
                'G' + str(row + 13) + ':L' + str(row + 14), 
                'प्राप्तकर्त्याची स्वाक्षरी', 
                sign
            )

            ###########################################################

            # sr_no       = excel_data[i + 4][0]
            # name        = excel_data[i + 4][1]

            # # M1:N1
            # sheet.merge_range(
            #     'M' + str(row + 1) + ':N' + str(row + 1), 
            #     'क्र: ' + str(sr_no), 
            #     data
            # )

            # # Q1:R1
            # sheet.merge_range(
            #     'Q' + str(row + 1) + ':R' + str(row + 1), 
            #     'फोन: 9371079745', 
            #     data
            # )

            # # O1:P1
            # sheet.merge_range(
            #     'O' + str(row + 1) + ':P' + str(row + 1), 
            #     '', 
            #     merge_format)

            # # M2:R2
            # sheet.merge_range(
            #     'M' + str(row + 2) + ':R' + str(row + 2), 
            #     'मातृछाया दुग्धालय\n 210, कसबा पेठ, पुणे-411011', 
            #     heading)

            # # M3:R3
            # sheet.merge_range(
            #     'M' + str(row + 3) + ':R' + str(row + 3), 
            #     name, 
            #     custom_font)

            # # M5:O5
            # sheet.merge_range(
            #     'M' + str(row + 5) + ':O' + str(row + 5), 
            #     'तपशील', 
            #     heading
            # )

            # # P5
            # sheet.write(
            #     'P'  + str(row + 5), 
            #     'प्रमाण', 
            #     heading
            # )

            # # Q5
            # sheet.write(
            #     'Q' + str(row + 5) , 
            #     'दर', 
            #     heading
            # )

            # # R5
            # sheet.write(
            #     'R' + str(row + 5), 
            #     'रक्कम', 
            #     heading
            # )

            # # M4:O4
            # sheet.merge_range(
            #     'M' + str(row + 4) + ':O' + str(row + 4), 
            #     datetime.datetime.strptime(bill_for, "%Y-%m-%d %H:%M:%S").strftime("%B-%Y"), 
            #     merge_format)

            # # P4:R4
            # sheet.merge_range(
            #     'P' + str(row + 4) + ':R' + str(row + 4), 
            #     'तारीख: ' + str(datetime.datetime.strptime(bill_gen, "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")), 
            #     data)

            # quantity        = excel_data[i + 4][2]
            # dahi            = excel_data[i + 4][3]
            # ghee            = excel_data[i + 4][4]
            # milk_type       = excel_data[i + 4][5]
            # amount          = excel_data[i + 4][6]
            # previous_blnc   = excel_data[i + 4][7]
            # total           = excel_data[i + 4][8]

            # quantity1   = excel_data[i + 5][2]
            # milk_type1  = excel_data[i + 5][5]
            # amount1     = excel_data[i + 5][6]
            
            # # M6:O6
            # sheet.merge_range(
            #     'M' + str(row + 6) + ':O' + str(row + 6), 
            #     'म्हशींचे दूध', 
            #     data
            # )
            # if milk_type == 'M':
            #     sheet.write('P' + str(row + 6), round(float(quantity), 2), data)
            #     sheet.write('Q' + str(row + 6), str(m_rate) + '/Ltr', data)
            #     sheet.write('R' + str(row + 6), round(float(amount), 2), data)
            # elif milk_type1 == 'M':
            #     sheet.write('P' + str(row + 6), round(float(quantity1), 2), data)
            #     sheet.write('Q' + str(row + 6), str(m_rate) + '/Ltr', data)
            #     sheet.write('R' + str(row + 6), round(float(amount1), 2), data)
            # else:
            #     sheet.write_blank('P' + str(row + 6), '', data)
            #     sheet.write_blank('Q' + str(row + 6), '', data)
            #     sheet.write_blank('R' + str(row + 6), '', data)

            # # M7:O7
            # sheet.merge_range(
            #     'M' + str(row + 7) + ':O' + str(row + 7), 
            #     'गायीचे दूध', 
            #     data
            # )
            # if milk_type == 'G':
            #     sheet.write('P'  + str(row + 7), round(float(quantity), 2), data)
            #     sheet.write('Q' + str(row + 7), str(g_rate) + '/Ltr', data)
            #     sheet.write('R' + str(row + 7), round(float(amount), 2), data)
            # elif milk_type1 == 'G':
            #     sheet.write('P' + str(row + 7), round(float(quantity1), 2), data)
            #     sheet.write('Q' + str(row + 7), str(g_rate) + '/Ltr', data)
            #     sheet.write('R' + str(row + 7), round(float(amount1), 2), data)
            # else:
            #     sheet.write_blank('P'  + str(row + 7), '', data)
            #     sheet.write_blank('Q' + str(row + 7), '', data)
            #     sheet.write_blank('R' + str(row + 7), '', data)

            # # M8:O8
            # sheet.merge_range(
            #     'M' + str(row + 8) + ':O' + str(row + 8), 
            #     'दही', 
            #     data
            # )

            # # P8
            # sheet.write_blank('P' + str(row + 8), '', data)

            # # Q8
            # sheet.write('Q' + str(row + 8), str(dahi_rate) + '/Kg', data)

            # # R8
            # sheet.write('R' + str(row + 8), dahi, data)

            # # M9:O9
            # sheet.merge_range(
            #     'M' + str(row + 9) + ':O' + str(row + 9), 
            #     'तूप', 
            #     data
            # )

            # # P9
            # sheet.write_blank('P' + str(row + 9), '', data)

            # # Q9
            # sheet.write('Q' + str(row + 9), str(ghee_rate) + '/Kg', data)

            # # R9
            # sheet.write('R' + str(row + 9), ghee, data)
            # # sheet.merge_range('A9:C9', 'मागील शिल्लक', data)

            # # M10:O10
            # sheet.merge_range(
            #     'M' + str(row + 10) + ':O' + str(row + 10), 
            #     'मागील बाकी', 
            #     data
            # )

            # # M11:O11
            # sheet.merge_range(
            #     'M' + str(row + 11) + ':O' + str(row + 11), 
            #     'एकूण', 
            #     data
            # )
            
            # # P10
            # sheet.write_blank('P' + str(row + 10), '', data)

            # # Q10
            # sheet.write_blank('Q' + str(row + 10), '', data)

            # # R10
            # sheet.write('R' + str(row + 10), previous_blnc, data)

            # # R11
            # sheet.write('R' + str(row + 11), round(float(total), 2), data)

            # # M11:R12
            # sheet.merge_range(
            #     'M' + str(row + 12) + ':R' + str(row + 13), 
            #     'प्राप्तकर्त्याची स्वाक्षरी', 
            #     sign
            # )

            row += 14
        except IndexError:
            break
        
    book.close()
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=" + today.strftime('bills/%Y/%B/') + ".xlsx"

    output.close()
    messages.info(request, 'File generated successfully')
    request.session.clear()
    
    return response
